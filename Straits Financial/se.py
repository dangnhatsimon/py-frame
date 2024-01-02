import os
import pandas as pd
from tkinter import *
from tkinter import filedialog
from openpyxl import load_workbook

def search(folder_path, keyword):
    results = []
    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(folder_path, filename)
            df = pd.read_excel(file_path, header=None)
            for index, row in df.iterrows():
                for col_index, cell_value in enumerate(row):
                    if keyword.lower() in str(cell_value).lower():
                        result_info = {
                            "file": filename,
                            "row": index + 1,  # Excel rows are 1-indexed
                            "column": col_index + 1  # Excel columns are 1-indexed
                        }
                        results.append(result_info)
    return results

def display_results(results_listbox, results):
    results_listbox.delete(0, END)
    for result in results:
        result_str = f"File: {result['file']}, Row: {result['row']}, Column: {result['column']}"
        results_listbox.insert(END, result_str)

def open_file_and_highlight(file_path, row, column):
    wb = load_workbook(file_path)
    sheet = wb.active
    cell = sheet.cell(row=row, column=column)
    wb.active = sheet
    wb.save(file_path)
    os.system(f'start excel "{file_path}"')

def browse_folder(entry_widget):
    folder_path = filedialog.askdirectory()
    entry_widget.delete(0, END)
    entry_widget.insert(0, folder_path)

def on_result_click(results_listbox, results):
    selected_index = results_listbox.curselection()
    if selected_index:
        selected_result = results[selected_index[0]]
        open_file_and_highlight(os.path.join(folder_entry.get(), selected_result['file']), selected_result['row'], selected_result['column'])

def main():
    root = Tk()
    root.title('Excel Search Tool')

    folder_entry = Entry(root, width=50, bg='white', fg='black', borderwidth=5)
    folder_entry.insert(0, 'Enter your folder path: ')
    folder_entry.grid(row=0, column=0, padx=10, pady=10)

    browse_button = Button(root, text='Browse', padx=50, pady=10, command=lambda: browse_folder(folder_entry), fg='blue', bg='#ffffff')
    browse_button.grid(row=0, column=1, padx=10, pady=10)

    keyword_entry = Entry(root, width=50, bg='white', fg='black', borderwidth=5)
    keyword_entry.insert(0, 'Enter keyword: ')
    keyword_entry.grid(row=1, column=0, padx=10, pady=10)

    search_button = Button(root, text='Search', command=lambda: display_results(results_listbox, search(folder_entry.get(), keyword_entry.get())))
    search_button.grid(row=1, column=1, padx=10, pady=10)

    results_listbox = Listbox(root, width=80, height=10)
    results_listbox.grid(row=2, column=0, columnspan=2, pady=10)
    results_listbox.bind("<Double-Button-1>", lambda event: on_result_click(results_listbox, results_list))

    root.mainloop()

if __name__ == "__main__":
    main()
